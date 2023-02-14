from openpyxl import Workbook,load_workbook



def int_to_letters(value:int) -> str:
    letter = ''

    if value <= 0:
        return ""

    alpha_len = ord('Z')-ord('A')+1

    letter_value = int(value/alpha_len  )
    rest_value   = int(value%alpha_len )

    if letter_value == 0 and rest_value:     
        return chr(ord('A') + rest_value - 1)

    elif letter_value <= alpha_len and rest_value == 0:
        if letter_value > 1:
            letter += chr(ord('A') + letter_value - 2)
        letter += "Z"

    elif letter_value <= alpha_len and rest_value:
        letter += chr(ord('A') + letter_value - 1)
        letter += int_to_letters( rest_value )

    elif rest_value == 0:
        if letter_value > 1:
            letter += int_to_letters( letter_value -1) # chr(ord('A') + letter_value - 2)
        letter += "Z"
    else:
        letter += int_to_letters( letter_value)
        letter += int_to_letters( rest_value )

    return letter


def readin_xlsx(xlsx_filename:str) -> dict:

    res = {}

    wb = Workbook()
    wb = load_workbook(xlsx_filename)

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    column_index = {}

    i = 1
    while True:

        value = ws[f'{int_to_letters(i)}1'].value
#        print(f'{int_to_letters(i)}1 -- > {value}')
        if value == None:
            break
        res[ value ] = []
        column_index[ i ] = value

        j = 2
        while True:
            cell_value = ws[f'{int_to_letters(i)}{j}'].value

            if cell_value == None:
                break
            res[ value ].append(cell_value)

            j += 1

        i += 1

    return res




def read_headers(xlsx_filename:str) -> dict:

    res = {}

    wb = Workbook()
    wb = load_workbook(xlsx_filename)


    for ws in wb.worksheets:
        ws_title = ws.title
        res[ ws_title ] =  {}
        i = 1
        ducks = 0
        while True:
            cell_name = f"{int_to_letters(i)}1"
            cell_value = ws[ cell_name ].value

            if cell_value in [None, '']:
                ducks += 1
            else:
                ducks = 0
                res[ ws_title ][ cell_name] = cell_value 
            
            if ducks >= 5:
                break

            i += 1

        if res[ ws_title ] ==  {}:
            del res[ ws_title ] 

    return res
